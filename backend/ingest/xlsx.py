from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def read_xlsx_sheet_rows(path: str | Path, sheet_name: str) -> list[list[Any]]:
    workbook = load_workbook(Path(path), read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        available = ", ".join(workbook.sheetnames)
        raise ValueError(f"Sheet '{sheet_name}' was not found. Available sheets: {available}")
    worksheet = workbook[sheet_name]
    return [list(row) for row in worksheet.iter_rows(values_only=True)]


def dict_rows_from_header_sheet(path: str | Path, sheet_name: str, header_row: int = 1) -> list[dict[str, Any]]:
    rows = read_xlsx_sheet_rows(path, sheet_name)
    if header_row < 1 or header_row > len(rows):
        raise ValueError(f"Header row {header_row} is outside sheet '{sheet_name}'.")
    headers = [_normalize_header(value) for value in rows[header_row - 1]]
    dict_rows: list[dict[str, Any]] = []
    for row in rows[header_row:]:
        record = {
            header: row[index] if index < len(row) else None
            for index, header in enumerate(headers)
            if header
        }
        if any(value not in (None, "") for value in record.values()):
            dict_rows.append(record)
    return dict_rows


def _normalize_header(value: Any) -> str:
    return str(value or "").strip().lower().replace("\n", " ").replace(" ", "_")
