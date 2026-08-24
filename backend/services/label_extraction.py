from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from io import BytesIO
from pathlib import Path
from typing import Any, Literal

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, Field, model_validator
from sqlalchemy import and_, or_, select
from sqlalchemy.orm import Session, aliased

from backend.core.config import DEFAULT_PROJECT_UID
from backend.persistence.postgresql import models as db
from backend.persistence.postgresql.filter_presets import FilterPayload, validate_filter_payload


DEFAULT_LABEL_EXTRACTION_CONFIG_PATH = Path("data/label_extraction_configs.json")
LabelScopeType = Literal["all", "filter", "filter_preset", "entity_group", "source_cable_group", "cabinet_pair"]
LabelOutputMode = Literal["single_sheet", "category_sheets"]


class CabinetPairScope(BaseModel):
    source_cabinet_uid: str
    target_cabinet_uid: str
    include_reverse: bool = True


class LabelExtractionScope(BaseModel):
    source: LabelScopeType = "all"
    filter_payload: dict[str, Any] | None = None
    filter_preset_uid: str | None = None
    entity_group_uids: list[str] = Field(default_factory=list)
    cable_groups: list[str] = Field(default_factory=list)
    cabinet_pairs: list[CabinetPairScope] = Field(default_factory=list)


class LabelFieldConfig(BaseModel):
    a_label_source: Literal["stored_label", "port_uid", "port_name"] = "stored_label"
    z_label_source: Literal["stored_label", "port_uid", "port_name"] = "stored_label"


class LabelOutputConfig(BaseModel):
    mode: LabelOutputMode = "category_sheets"
    category_fields: list[Literal["pair", "cable_type", "group", "status", "construction_phase"]] = Field(
        default_factory=lambda: ["pair", "cable_type"]
    )


class LabelExtractionConfig(BaseModel):
    uid: str | None = None
    project_uid: str = DEFAULT_PROJECT_UID
    name: str
    description: str = ""
    scope: LabelExtractionScope = Field(default_factory=LabelExtractionScope)
    cable_filter_payload: dict[str, Any] | None = None
    labels: LabelFieldConfig = Field(default_factory=LabelFieldConfig)
    output: LabelOutputConfig = Field(default_factory=LabelOutputConfig)

    @model_validator(mode="after")
    def validate_scope_payloads(self) -> "LabelExtractionConfig":
        validate_label_extraction_config(self)
        return self


class LabelCableRow(BaseModel):
    uid: str
    a_port_uid: str
    z_port_uid: str
    a_cabinet_uid: str
    z_cabinet_uid: str
    a_port_name: str = ""
    z_port_name: str = ""
    cable_type: str = ""
    group: str = ""
    status: str = ""
    construction_phase: str = ""
    a_label: str = ""
    z_label: str = ""
    pair: str = ""
    category: str = ""


class DuplicateLabelRecord(BaseModel):
    label_text: str
    side: Literal["a", "z"]
    cable_uids: list[str]


class LabelExtractionSummary(BaseModel):
    total_cables: int
    counts_by_pair: dict[str, int] = Field(default_factory=dict)
    counts_by_category: dict[str, int] = Field(default_factory=dict)
    duplicate_labels: list[DuplicateLabelRecord] = Field(default_factory=list)
    empty_label_cable_uids: list[str] = Field(default_factory=list)


class LabelExtractionPreview(BaseModel):
    config: LabelExtractionConfig
    summary: LabelExtractionSummary
    sample_rows: list[LabelCableRow]


def validate_label_extraction_config(config: LabelExtractionConfig) -> None:
    scope = config.scope
    if scope.source == "filter" and not scope.filter_payload:
        raise ValueError("scope.filter_payload is required when scope.source is 'filter'.")
    if scope.source == "filter_preset" and not scope.filter_preset_uid:
        raise ValueError("scope.filter_preset_uid is required when scope.source is 'filter_preset'.")
    if scope.source == "entity_group" and not scope.entity_group_uids:
        raise ValueError("scope.entity_group_uids is required when scope.source is 'entity_group'.")
    if scope.source == "source_cable_group" and not scope.cable_groups:
        raise ValueError("scope.cable_groups is required when scope.source is 'source_cable_group'.")
    if scope.source == "cabinet_pair" and not scope.cabinet_pairs:
        raise ValueError("scope.cabinet_pairs is required when scope.source is 'cabinet_pair'.")
    if scope.filter_payload:
        validate_filter_payload("cable", scope.filter_payload)
    if config.cable_filter_payload:
        validate_filter_payload("cable", config.cable_filter_payload)


def built_in_label_extraction_configs(project_uid: str = DEFAULT_PROJECT_UID) -> list[LabelExtractionConfig]:
    return [
        LabelExtractionConfig(
            uid=f"{project_uid.lower()}-source-group-labels",
            project_uid=project_uid,
            name="Source cable group labels",
            description="Generate label sheets from one or more source GROUP values.",
            scope=LabelExtractionScope(source="source_cable_group", cable_groups=["GROUP NAME"]),
        ),
        LabelExtractionConfig(
            uid=f"{project_uid.lower()}-cabinet-pair-labels",
            project_uid=project_uid,
            name="Cabinet pair labels",
            description="Generate label sheets for explicit A/Z cabinet pairs.",
            scope=LabelExtractionScope(
                source="cabinet_pair",
                cabinet_pairs=[CabinetPairScope(source_cabinet_uid="DH1:001", target_cabinet_uid="DH1:002")],
            ),
        ),
    ]


def load_label_extraction_configs(path: str | Path = DEFAULT_LABEL_EXTRACTION_CONFIG_PATH) -> list[LabelExtractionConfig]:
    config_path = Path(path)
    if not config_path.exists():
        return []
    with config_path.open(encoding="utf-8") as file:
        payload = json.load(file)
    items = payload.get("configs", payload) if isinstance(payload, dict) else payload
    return [LabelExtractionConfig.model_validate(item) for item in items]


def save_label_extraction_config(config: LabelExtractionConfig, path: str | Path = DEFAULT_LABEL_EXTRACTION_CONFIG_PATH) -> LabelExtractionConfig:
    config_path = Path(path)
    config_path.parent.mkdir(parents=True, exist_ok=True)
    uid = config.uid or _config_uid(config.project_uid, config.name)
    saved_config = config.model_copy(update={"uid": uid})
    configs = load_label_extraction_configs(config_path)
    by_uid = {item.uid: item for item in configs if item.uid}
    by_uid[uid] = saved_config
    payload = {"configs": [_model_dump(item) for item in sorted(by_uid.values(), key=lambda item: (item.project_uid, item.name))]}
    with config_path.open("w", encoding="utf-8") as file:
        json.dump(payload, file, indent=2)
        file.write("\n")
    return saved_config


def resolve_label_extraction_config(
    uid: str,
    *,
    project_uid: str,
    path: str | Path = DEFAULT_LABEL_EXTRACTION_CONFIG_PATH,
) -> LabelExtractionConfig | None:
    candidates = [*load_label_extraction_configs(path), *built_in_label_extraction_configs(project_uid)]
    for config in candidates:
        if config.uid == uid and config.project_uid == project_uid:
            return config
    return None


def resolve_label_cable_rows(session: Session, config: LabelExtractionConfig, *, limit: int = 100_000) -> list[LabelCableRow]:
    a_port = aliased(db.Port)
    z_port = aliased(db.Port)
    statement = (
        select(
            db.Cable.uid,
            db.Cable.a_port_uid,
            db.Cable.z_port_uid,
            db.Cable.cable_type,
            db.Cable.cable_group,
            db.Cable.import_status,
            db.Cable.construction_phase,
            db.Cable.a_label_text,
            db.Cable.z_label_text,
            a_port.cabinet_uid.label("a_cabinet_uid"),
            z_port.cabinet_uid.label("z_cabinet_uid"),
            a_port.port_name.label("a_port_name"),
            z_port.port_name.label("z_port_name"),
        )
        .join(a_port, db.Cable.a_port_uid == a_port.uid)
        .join(z_port, db.Cable.z_port_uid == z_port.uid)
        .where(
            db.Cable.project_uid == config.project_uid,
            db.Cable.deleted_at.is_(None),
            a_port.deleted_at.is_(None),
            z_port.deleted_at.is_(None),
        )
    )
    clauses = _scope_clauses(session, config, a_port, z_port)
    if clauses:
        statement = statement.where(*clauses)
    if config.cable_filter_payload:
        statement = statement.where(_filter_payload_clause(validate_filter_payload("cable", config.cable_filter_payload)))
    rows = session.execute(
        statement.order_by(a_port.cabinet_uid, z_port.cabinet_uid, db.Cable.cable_type, db.Cable.uid).limit(limit)
    ).all()
    return [_label_row(row, config) for row in rows]


def preview_label_extraction_rows(config: LabelExtractionConfig, rows: list[LabelCableRow], *, sample_size: int = 25) -> LabelExtractionPreview:
    return LabelExtractionPreview(config=config, summary=summarize_label_rows(rows), sample_rows=rows[:sample_size])


def summarize_label_rows(rows: list[LabelCableRow]) -> LabelExtractionSummary:
    a_labels: dict[str, list[str]] = defaultdict(list)
    z_labels: dict[str, list[str]] = defaultdict(list)
    empty_label_cable_uids: list[str] = []
    for row in rows:
        if row.a_label:
            a_labels[row.a_label].append(row.uid)
        if row.z_label:
            z_labels[row.z_label].append(row.uid)
        if not row.a_label or not row.z_label:
            empty_label_cable_uids.append(row.uid)
    duplicates = [
        DuplicateLabelRecord(label_text=label, side=side, cable_uids=uids)
        for side, labels in (("a", a_labels), ("z", z_labels))
        for label, uids in labels.items()
        if len(uids) > 1
    ]
    duplicates.sort(key=lambda item: (item.side, item.label_text))
    return LabelExtractionSummary(
        total_cables=len(rows),
        counts_by_pair=dict(sorted(Counter(row.pair for row in rows).items())),
        counts_by_category=dict(sorted(Counter(row.category for row in rows).items())),
        duplicate_labels=duplicates,
        empty_label_cable_uids=empty_label_cable_uids,
    )


def build_label_workbook_bytes(config: LabelExtractionConfig, rows: list[LabelCableRow]) -> bytes:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    _write_summary_sheet(summary_sheet, config, summarize_label_rows(rows))

    grouped_rows: dict[str, list[LabelCableRow]] = {"Labels": rows}
    if config.output.mode == "category_sheets":
        grouped = defaultdict(list)
        for row in rows:
            grouped[row.category or "Uncategorized"].append(row)
        grouped_rows = dict(sorted(grouped.items()))
    for sheet_name, sheet_rows in grouped_rows.items():
        worksheet = workbook.create_sheet(_sheet_title(sheet_name))
        _write_label_sheet(worksheet, sheet_rows)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _scope_clauses(session: Session, config: LabelExtractionConfig, a_port: Any, z_port: Any) -> list[Any]:
    scope = config.scope
    if scope.source == "all":
        return []
    if scope.source == "filter" and scope.filter_payload:
        return [_filter_payload_clause(validate_filter_payload("cable", scope.filter_payload))]
    if scope.source == "filter_preset" and scope.filter_preset_uid:
        preset = session.get(db.FilterPreset, scope.filter_preset_uid)
        if preset is None or preset.deleted_at is not None or preset.project_uid != config.project_uid or preset.entity_type != "cable":
            raise ValueError("Filter preset was not found for this project.")
        return [_filter_payload_clause(validate_filter_payload("cable", preset.filter_payload))]
    if scope.source == "entity_group":
        return [
            db.Cable.uid.in_(
                select(db.EntityGroupMember.entity_uid).where(
                    db.EntityGroupMember.group_uid.in_(scope.entity_group_uids),
                    db.EntityGroupMember.entity_type == "cable",
                )
            )
        ]
    if scope.source == "source_cable_group":
        return [db.Cable.cable_group.in_(scope.cable_groups)]
    if scope.source == "cabinet_pair":
        pair_clauses = []
        for pair in scope.cabinet_pairs:
            source_uid = pair.source_cabinet_uid.upper()
            target_uid = pair.target_cabinet_uid.upper()
            forward = and_(a_port.cabinet_uid == source_uid, z_port.cabinet_uid == target_uid)
            if pair.include_reverse:
                pair_clauses.append(or_(forward, and_(a_port.cabinet_uid == target_uid, z_port.cabinet_uid == source_uid)))
            else:
                pair_clauses.append(forward)
        return [or_(*pair_clauses)] if pair_clauses else []
    return []


def _filter_payload_clause(filter_payload: FilterPayload) -> Any:
    clauses = [_filter_rule_clause(rule) for rule in filter_payload.rules]
    if not clauses:
        return True
    return or_(*clauses) if filter_payload.logic == "or" else and_(*clauses)


def _filter_rule_clause(rule: Any) -> Any:
    column = _filter_column(rule.field)
    operator = rule.operator.strip().lower()
    value = rule.value
    if operator == "equals":
        return column == value
    if operator == "not_equals":
        return column != value
    if operator == "in":
        return column.in_(value)
    if operator == "not_in":
        return column.notin_(value)
    if operator == "contains":
        return column.ilike(f"%{value}%")
    if operator == "starts_with":
        return column.ilike(f"{value}%")
    if operator == "between":
        return column.between(value[0], value[1])
    if operator == "gt":
        return column > value
    if operator == "gte":
        return column >= value
    if operator == "lt":
        return column < value
    if operator == "lte":
        return column <= value
    if operator == "is_blank":
        return or_(column.is_(None), column == "")
    if operator == "is_not_blank":
        return and_(column.is_not(None), column != "")
    raise ValueError(f"Unsupported filter operator '{rule.operator}'.")


def _filter_column(field: str) -> Any:
    field_map = {"status": "import_status", "group": "cable_group"}
    return getattr(db.Cable, field_map.get(field, field))


def _label_row(row: Any, config: LabelExtractionConfig) -> LabelCableRow:
    pair = f"{row.a_cabinet_uid} -> {row.z_cabinet_uid}"
    base = LabelCableRow(
        uid=row.uid,
        a_port_uid=row.a_port_uid,
        z_port_uid=row.z_port_uid,
        a_cabinet_uid=row.a_cabinet_uid,
        z_cabinet_uid=row.z_cabinet_uid,
        a_port_name=row.a_port_name or "",
        z_port_name=row.z_port_name or "",
        cable_type=row.cable_type or "",
        group=row.cable_group or "",
        status=row.import_status or "",
        construction_phase=row.construction_phase or "",
        a_label=_label_value(config.labels.a_label_source, row.a_label_text, row.a_port_uid, row.a_port_name),
        z_label=_label_value(config.labels.z_label_source, row.z_label_text, row.z_port_uid, row.z_port_name),
        pair=pair,
    )
    return base.model_copy(update={"category": _category(base, config.output.category_fields)})


def _label_value(source: str, stored_label: str, port_uid: str, port_name: str) -> str:
    if source == "port_uid":
        return port_uid or ""
    if source == "port_name":
        return port_name or ""
    return stored_label or ""


def _category(row: LabelCableRow, fields: list[str]) -> str:
    values = [str(getattr(row, field) or "Unspecified") for field in fields]
    return " / ".join(values) if values else "Labels"


def _write_summary_sheet(sheet: Any, config: LabelExtractionConfig, summary: LabelExtractionSummary) -> None:
    rows = [
        ("Config", config.name),
        ("Project", config.project_uid),
        ("Total cables", summary.total_cables),
        ("Duplicate labels", len(summary.duplicate_labels)),
        ("Rows with empty labels", len(summary.empty_label_cable_uids)),
        ("", ""),
        ("Category", "Cable count"),
    ]
    for category, count in summary.counts_by_category.items():
        rows.append((category, count))
    for row in rows:
        sheet.append(row)
    _style_header_row(sheet, 7)
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 18


def _write_label_sheet(sheet: Any, rows: list[LabelCableRow]) -> None:
    headers = ["Cable UID", "Pair", "Cable Type", "Group", "Status", "A Cabinet", "A Port", "A Label", "Z Cabinet", "Z Port", "Z Label"]
    sheet.append(headers)
    _style_header_row(sheet, 1)
    for row in rows:
        sheet.append(
            [
                row.uid,
                row.pair,
                row.cable_type,
                row.group,
                row.status,
                row.a_cabinet_uid,
                row.a_port_name,
                row.a_label,
                row.z_cabinet_uid,
                row.z_port_name,
                row.z_label,
            ]
        )
    widths = [18, 30, 16, 24, 26, 16, 18, 32, 16, 18, 32]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width


def _style_header_row(sheet: Any, row_number: int) -> None:
    fill = PatternFill("solid", fgColor="1F2937")
    for cell in sheet[row_number]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill


def _sheet_title(value: str) -> str:
    title = re.sub(r"[\[\]:*?/\\]", "-", value).strip() or "Labels"
    return title[:31]


def _config_uid(project_uid: str, name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", name.casefold()).strip("-") or "label-config"
    return f"{project_uid.casefold()}-{slug}"


def _model_dump(model: BaseModel) -> dict[str, Any]:
    if hasattr(model, "model_dump"):
        return model.model_dump()
    return model.dict()
