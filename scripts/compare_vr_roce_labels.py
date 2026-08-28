from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from backend.ingest.cleaners.lbb01 import (  # noqa: E402
    DEFAULT_VR_ROCE_SHEETS,
    _normalize_lbb_non_roce_row,
)
from backend.ingest.cutsheet import ingest_cutsheet_rows  # noqa: E402
from backend.ingest.ods import read_ods_sheet_rows  # noqa: E402


RACK_MIN = 1141
RACK_MAX = 1160


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--old", required=True, type=Path)
    parser.add_argument("--new", required=True, type=Path)
    parser.add_argument("--out-dir", default=Path("tmp/vr_roce_compare"), type=Path)
    args = parser.parse_args()

    args.out_dir.mkdir(parents=True, exist_ok=True)
    old_rows = read_workbook_rows(args.old)
    new_rows = read_workbook_rows(args.new)

    old_records = comparable_records(old_rows)
    new_records = comparable_records(new_rows)
    old_by_key = {record["key"]: record for record in old_records}
    new_by_key = {record["key"]: record for record in new_records}

    old_keys = set(old_by_key)
    new_keys = set(new_by_key)
    added = [new_by_key[key] for key in sorted(new_keys - old_keys)]
    removed = [old_by_key[key] for key in sorted(old_keys - new_keys)]
    changed = []
    for key in sorted(old_keys & new_keys):
        old_record = old_by_key[key]
        new_record = new_by_key[key]
        deltas = {
            field: {"old": old_record.get(field, ""), "new": new_record.get(field, "")}
            for field in [
                "status",
                "cable",
                "a_side_dns_name",
                "a_loc_cab_ru",
                "a_model",
                "a_mpo",
                "a_port",
                "a_connector",
                "a_interface",
                "a_optic",
                "z_side_dns_name",
                "z_loc_cab_ru",
                "z_model",
                "z_mpo",
                "z_port",
                "z_connector",
                "z_interface",
                "z_optic",
            ]
            if old_record.get(field, "") != new_record.get(field, "")
        }
        if deltas:
            changed.append({"key": key, "sheet": new_record["sheet"], "deltas": deltas})

    write_csv(args.out_dir / "added.csv", added)
    write_csv(args.out_dir / "removed.csv", removed)
    write_json(args.out_dir / "changed.json", changed)

    summary = {
        "old_rows": len(old_records),
        "new_rows": len(new_records),
        "added": len(added),
        "removed": len(removed),
        "changed_same_endpoint": len(changed),
        "old_rows_by_sheet": dict(Counter(record["sheet"] for record in old_records)),
        "new_rows_by_sheet": dict(Counter(record["sheet"] for record in new_records)),
        "added_by_sheet": dict(Counter(record["sheet"] for record in added)),
        "removed_by_sheet": dict(Counter(record["sheet"] for record in removed)),
        "changed_by_sheet": dict(Counter(record["sheet"] for record in changed)),
        "affected_racks": summarize_by_rack(added, removed, changed, old_by_key, new_by_key),
    }
    write_json(args.out_dir / "summary.json", summary)
    print(json.dumps(summary, indent=2))


def read_workbook_rows(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".ods":
        rows_by_sheet = {
            sheet_name: matrix_to_dicts(read_ods_sheet_rows(path, sheet_name))
            for sheet_name in DEFAULT_VR_ROCE_SHEETS
        }
    else:
        workbook = load_workbook(path, read_only=True, data_only=True)
        rows_by_sheet = {
            sheet_name: worksheet_to_dicts(workbook[sheet_name])
            for sheet_name in DEFAULT_VR_ROCE_SHEETS
            if sheet_name in workbook.sheetnames
        }

    records = []
    for sheet_name, rows in rows_by_sheet.items():
        for row_number, row in enumerate(rows, start=2):
            normalized = _normalize_lbb_non_roce_row(row)
            if any(str(value or "").strip() for value in normalized.values()):
                normalized["sheet"] = sheet_name
                normalized["source_row"] = row_number
                records.append(normalized)
    return records


def matrix_to_dicts(rows: list[list[str]]) -> list[dict[str, Any]]:
    if not rows:
        return []
    headers = rows[0]
    return [
        {headers[index]: row[index] if index < len(row) else "" for index in range(len(headers))}
        for row in rows[1:]
    ]


def worksheet_to_dicts(worksheet: Any) -> list[dict[str, Any]]:
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    records = []
    for row in rows[1:]:
        record = {
            headers[index]: row[index] if index < len(row) else ""
            for index in range(len(headers))
            if headers[index]
        }
        if any(value not in (None, "") for value in record.values()):
            records.append(record)
    return records


def comparable_records(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    result = ingest_cutsheet_rows(rows)
    by_endpoint = {
        (parsed.a_port_uid, parsed.z_port_uid): parsed
        for parsed in result.rows
        if rack_in_scope(parsed.a_cabinet_id) or rack_in_scope(parsed.z_cabinet_id)
    }
    records = []
    for row in rows:
        a_loc = str(row.get("a_loc_cab_ru", ""))
        z_loc = str(row.get("z_loc_cab_ru", ""))
        if not (loc_rack_in_scope(a_loc) or loc_rack_in_scope(z_loc)):
            continue
        parsed = by_endpoint.get((port_uid(a_loc, row.get("a_port") or row.get("a_interface")), port_uid(z_loc, row.get("z_port") or row.get("z_interface"))))
        key = endpoint_key(row)
        record = dict(row)
        record["key"] = key
        record["parsed_a_port_uid"] = parsed.a_port_uid if parsed else ""
        record["parsed_z_port_uid"] = parsed.z_port_uid if parsed else ""
        records.append(record)
    return records


def endpoint_key(row: dict[str, Any]) -> str:
    return "|".join(
        [
            str(row.get("sheet", "")),
            str(row.get("a_loc_cab_ru", "")),
            str(row.get("a_port") or row.get("a_interface") or ""),
            str(row.get("z_loc_cab_ru", "")),
            str(row.get("z_port") or row.get("z_interface") or ""),
        ]
    )


def port_uid(loc: str, port: Any) -> str:
    parts = str(loc).split(":")
    if len(parts) != 3:
        return ""
    return f"{parts[0].upper()}:{parts[1].zfill(3)}:{parts[2]}:{str(port or '').strip()}"


def loc_rack_in_scope(loc: str) -> bool:
    parts = str(loc).split(":")
    return len(parts) == 3 and rack_in_scope(parts[1])


def rack_in_scope(cabinet_id: str) -> bool:
    return str(cabinet_id).isdigit() and RACK_MIN <= int(cabinet_id) <= RACK_MAX


def summarize_by_rack(
    added: list[dict[str, Any]],
    removed: list[dict[str, Any]],
    changed: list[dict[str, Any]],
    old_by_key: dict[str, dict[str, Any]],
    new_by_key: dict[str, dict[str, Any]],
) -> dict[str, dict[str, int]]:
    rack_counts: dict[str, Counter[str]] = defaultdict(Counter)
    for label, records in (("added", added), ("removed", removed)):
        for record in records:
            for rack in row_racks(record):
                rack_counts[rack][label] += 1
    for item in changed:
        records = [old_by_key.get(item["key"], {}), new_by_key.get(item["key"], {})]
        for record in records:
            for rack in row_racks(record):
                rack_counts[rack]["changed"] += 1
    return {rack: dict(counts) for rack, counts in sorted(rack_counts.items())}


def row_racks(row: dict[str, Any]) -> set[str]:
    racks = set()
    for key in ("a_loc_cab_ru", "z_loc_cab_ru"):
        parts = str(row.get(key, "")).split(":")
        if len(parts) == 3 and rack_in_scope(parts[1]):
            racks.add(str(int(parts[1])))
    return racks


def write_csv(path: Path, records: list[dict[str, Any]]) -> None:
    fields = sorted({field for record in records for field in record})
    with path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.DictWriter(file, fieldnames=fields)
        writer.writeheader()
        writer.writerows(records)


def write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main()
