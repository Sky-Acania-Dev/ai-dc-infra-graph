import unittest
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from backend.services.label_extraction import (
    LabelCableRow,
    LabelExtractionConfig,
    LabelExtractionScope,
    build_label_workbook_bytes,
    load_label_extraction_configs,
    preview_label_extraction_rows,
    save_label_extraction_config,
    summarize_label_rows,
)


class LabelExtractionTests(unittest.TestCase):
    def test_validates_scope_specific_requirements(self) -> None:
        with self.assertRaisesRegex(ValueError, "scope.cable_groups is required"):
            LabelExtractionConfig(
                project_uid="LBB01",
                name="Missing source group",
                scope=LabelExtractionScope(source="source_cable_group"),
            )

    def test_summarizes_duplicate_and_empty_labels(self) -> None:
        rows = [
            _row("CBL-1", "DH1:001 -> DH1:002", "LC", "A-1", "Z-1"),
            _row("CBL-2", "DH1:001 -> DH1:002", "LC", "A-1", ""),
            _row("CBL-3", "DH1:003 -> DH1:004", "MPO", "A-3", "Z-3"),
        ]

        summary = summarize_label_rows(rows)

        self.assertEqual(summary.total_cables, 3)
        self.assertEqual(summary.counts_by_pair["DH1:001 -> DH1:002"], 2)
        self.assertEqual(summary.counts_by_category["DH1:001 -> DH1:002 / LC"], 2)
        self.assertEqual(summary.duplicate_labels[0].label_text, "A-1")
        self.assertEqual(summary.empty_label_cable_uids, ["CBL-2"])

    def test_builds_summary_and_category_sheets(self) -> None:
        config = LabelExtractionConfig(
            uid="lbb01-labels",
            project_uid="LBB01",
            name="LBB01 Labels",
            scope=LabelExtractionScope(source="source_cable_group", cable_groups=["RO"]),
        )
        rows = [
            _row("CBL-1", "DH1:001 -> DH1:002", "LC", "A-1", "Z-1"),
            _row("CBL-2", "DH1:001 -> DH1:002", "LC", "A-2", "Z-2"),
            _row("CBL-3", "DH1:003 -> DH1:004", "MPO", "A-3", "Z-3"),
        ]

        workbook = load_workbook(BytesIO(build_label_workbook_bytes(config, rows)))

        self.assertEqual(workbook.sheetnames, ["Summary", "DH1-001 -> DH1-002 - LC", "DH1-003 -> DH1-004 - MPO"])
        self.assertEqual(workbook["Summary"]["B3"].value, 3)
        self.assertEqual(workbook["DH1-001 -> DH1-002 - LC"].max_row, 3)

    def test_saves_and_loads_file_backed_configs(self) -> None:
        config = LabelExtractionConfig(
            project_uid="LBB01",
            name="RO labels",
            scope=LabelExtractionScope(source="source_cable_group", cable_groups=["RO"]),
        )
        path = Path("data/runtime/test_label_extraction_configs.json")
        if path.exists():
            path.unlink()
        try:
            saved = save_label_extraction_config(config, path)
            loaded = load_label_extraction_configs(path)
        finally:
            if path.exists():
                path.unlink()

        self.assertEqual(saved.uid, "lbb01-ro-labels")
        self.assertEqual(len(loaded), 1)
        self.assertEqual(loaded[0].scope.cable_groups, ["RO"])

    def test_preview_limits_sample_rows(self) -> None:
        config = LabelExtractionConfig(name="All labels")
        rows = [_row(f"CBL-{index}", "DH1:001 -> DH1:002", "LC", f"A-{index}", f"Z-{index}") for index in range(5)]

        preview = preview_label_extraction_rows(config, rows, sample_size=2)

        self.assertEqual(preview.summary.total_cables, 5)
        self.assertEqual([row.uid for row in preview.sample_rows], ["CBL-0", "CBL-1"])


def _row(uid: str, pair: str, cable_type: str, a_label: str, z_label: str) -> LabelCableRow:
    a_cabinet, z_cabinet = pair.split(" -> ")
    row = LabelCableRow(
        uid=uid,
        a_port_uid=f"{a_cabinet}:1:p1",
        z_port_uid=f"{z_cabinet}:1:p1",
        a_cabinet_uid=a_cabinet,
        z_cabinet_uid=z_cabinet,
        a_port_name="p1",
        z_port_name="p1",
        cable_type=cable_type,
        group="RO",
        status="Cable Is Ran: Complete",
        construction_phase="Backbone",
        a_label=a_label,
        z_label=z_label,
        pair=pair,
    )
    return row.model_copy(update={"category": f"{pair} / {cable_type}"})


if __name__ == "__main__":
    unittest.main()
